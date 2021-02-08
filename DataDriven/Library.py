import json
import jsonpath
import requests
import openpyxl


class Common:

    def __init__(self, FileNamePath, SheetName):
        global wb, sheet
        wb = openpyxl.load_workbook(FileNamePath)
        sheet = wb[SheetName]


    def fetch_row_count(self):
        rows = sheet.max_row
        return rows

    def fetch_column_count(self):
        col = sheet.max_column

    def fetch_key_name(self):
        c = sheet.max_column
        li = []
        for i in range(1, c+1):
            cell = sheet.cell(row=i, column=i)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sheet.max_column
        for i in range (1, c+1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value
            return jsonRequest